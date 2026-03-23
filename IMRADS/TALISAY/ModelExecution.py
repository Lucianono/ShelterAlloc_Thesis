import openpyxl
from openpyxl.styles import Font
from WORKModel_Penalized import WORKModel
from BNSTModel_Penalized import BNSTModel
from BNTModel_Penalized import BNTModel
from BSTModel_Penalized import BSTModel


def save_to_excel(all_results, path="ga_results.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    headers = ["#", "Fitness", "Gen Last Updated", "Runtime", "Allocation"]

    for model_name, results in all_results.items():
        ws = wb.create_sheet(title=model_name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for i, r in enumerate(results, start=1):
            ws.append([
                i,
                str(r["fitness"]),
                r["gen_last_updated"],
                r["runtime"],
                str(r["allocation"]),
            ])

    wb.save(path)
    print(f"\nResults saved to: {path}")


MODELS = {
    "WORK":  WORKModel,
    "BNST":  BNSTModel,
    "BNT":   BNTModel,
    "BST":   BSTModel,
}

NUM_RUNS = 10
all_results = {}

for model_name, ModelClass in MODELS.items():
    print(f"\n{'='*50}")
    print(f"  MODEL: {model_name}")
    print(f"{'='*50}")
    results = []

    for run in range(1, NUM_RUNS + 1):
        print(f"\n--- {model_name} | Run {run} / {NUM_RUNS} ---")
        model = ModelClass()
        result = model.start_model()
        results.append(result)
        print(f"fitness={result['fitness']} | gen_last={result['gen_last_updated']} | time={result['runtime']}")

    all_results[model_name] = results

save_to_excel(all_results, path="ga_results.xlsx")