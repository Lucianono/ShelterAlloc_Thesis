import sys
from PySide6.QtWidgets import QPushButton, QCheckBox, QDialog, QLabel, QMessageBox, QInputDialog, QFileDialog, QTableWidgetItem, QWidget, QHBoxLayout, QTextEdit,QLineEdit
from PySide6.QtGui import QIcon, QCursor
from PySide6.QtCore import Qt, QUrl, QThread
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWebEngineCore import QWebEngineSettings
from ui_shelterallocationreport import Ui_ShelterAllocationReport
import pandas as pd
import os
import msoffcrypto
from io import BytesIO
from functools import partial
from pathfinder import PathfindingWorker
from optimizedRouting import run_optimization
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

class ShelterAllocationReport(QDialog):
    def __init__(self,show_Save_btn = True):
        super().__init__()  # Initialize the QDialog (or QWidget)
        self.ui = Ui_ShelterAllocationReport()  # Create an instance of the UI class
        self.ui.setupUi(self)  # Set up the UI on the current widget (QDialog)
        self.setModal(True)
        self.setWindowTitle("Shelter Allocation Report")
        self.save_dir = os.path.join(os.path.expanduser("~"), "Documents", "SLASystem")
        self.setWindowIcon(QIcon(os.path.join(sys._MEIPASS, "ICONS", "logo.png")))
        self.setAttribute(Qt.WA_DeleteOnClose)

        self.map_path = os.path.join(self.save_dir, "optimized-routes-map.html")
        self.ui.webEngineView.setUrl(QUrl.fromLocalFile(self.map_path))

        self.ui.webEngineView.settings().setAttribute(QWebEngineSettings.LocalContentCanAccessRemoteUrls, True)
        self.ui.webEngineView.settings().setAttribute(QWebEngineSettings.JavascriptEnabled, True)

        self.ui.pushButton_2.clicked.connect(self.save_report)
        self.ui.pushButton_3.clicked.connect(self.show_more_details)
        self.ui.pushButton.clicked.connect(self.close)

        self.load_table_data("allocation_results.xlsx")
        self.reload_label()

        if not show_Save_btn:
            self.ui.pushButton_2.hide()

    def save_report(self):


        # Save DataFrame to an in-memory buffer
        buffer = BytesIO()

        # Load the Excel template
        template_path = "ReportTemplate.xlsx"
        wb = load_workbook(os.path.join(sys._MEIPASS,template_path))
        ws1 = wb["Shelter Location-Allocation"]
        ws2 = wb["Report Analysis"]
        ws3 = wb["Community Data"]
        ws4 = wb["Shelter Data"]

        # 1st sheet
        df = pd.read_excel(os.path.join(self.save_dir,"allocation_results.xlsx"),header=0)
        df2 = pd.read_excel(os.path.join(self.save_dir,"modelCommData.xlsx"),usecols=["Name","Latitude","Longitude"],header=0)
        df3 = pd.read_excel(os.path.join(self.save_dir,"modelShelData.xlsx"),usecols=["Name","Latitude","Longitude"],header=0)
        merged_df = df.merge(df2, left_on="Community", right_on="Name", how="inner").merge(df3, left_on="Shelter Assigned", right_on="Name", how="inner")
        selected_df = merged_df[["Community", "Allocated Population" , "Latitude_x", "Longitude_y", "Shelter Assigned", "Level", "Latitude_y", "Longitude_y"]]

        start_row = 5
        for row_idx, row_data in selected_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                ws1.cell(row=row_idx+start_row, column=col_idx, value=value)

        # 2nd sheet
        model_results_path = os.path.join(self.save_dir, "modelPerformanceResult.txt")
        with open(model_results_path, "r") as file:
            content = file.readlines()
        content.reverse()

        start_row = 2
        for row_data in content:
            row_data = row_data.replace("\n", "")
            ws2.insert_rows(start_row)

            if "---" in row_data:
                ws2.cell(row=start_row, column=1).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid") 
                ws2.cell(row=start_row, column=1).font = Font(color="ffffff", bold=True)

            ws2.cell(row=start_row, column=1, value=row_data)

        # 3rd sheet
        df = pd.read_excel(os.path.join(self.save_dir,"modelCommData.xlsx"),header=0)
        start_row = 2
        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                ws3.cell(row=row_idx+start_row, column=col_idx, value=value)

        # 4th sheet
        df = pd.read_excel(os.path.join(self.save_dir,"modelShelData.xlsx"),header=0)
        start_row = 2
        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                ws4.cell(row=row_idx+start_row, column=col_idx, value=value)

        
        ws1.protection.sheet = True
        ws2.protection.sheet = True
        ws3.protection.sheet = True
        ws4.protection.sheet = True

        # Save as a new file (preserves original template formatting)
        wb.save(buffer)
        buffer.seek(0)  # Reset buffer position

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self, 
            "Save File", "", "All Files(*);;Excel Files(*.xlsx)", options = options)
        
        if fileName :
            if not fileName.endswith(".xlsx"):
                fileName += ".xlsx"
                
            response = QMessageBox.question(self, "Save Report", "Do you want to protect this report?", QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:

                text, ok = QInputDialog().getText(self, "Protect report","Enter password:",QLineEdit.Password)
                if ok and text:
                    password = text
                    # Encrypt and save the file
                    with open(f"{fileName}.xlsx", "wb") as f_out:
                        encryptor = msoffcrypto.OfficeFile(buffer)
                        encryptor.encrypt( password, f_out)
            else :
                with open(f"{fileName}.xlsx", "wb") as f:
                    f.write(buffer.getvalue())  # Write the buffer content to the file

            

        return

    def show_more_details(self):
        model_results_path = os.path.join(self.save_dir, "modelPerformanceResult.txt")
        with open(model_results_path, "r") as file:
            content = file.readlines()

        formatted_text = "<br>".join(line.strip() for line in content)
        QMessageBox.information(self, "Full Details", formatted_text)

    def reload_label(self):
        model_results_path = os.path.join(self.save_dir, "modelPerformanceResult.txt")
        with open(model_results_path, "r") as file:
            content = [next(file).strip() for _ in range(3)]

        formatted_text = "<br>".join(line.strip() for line in content)
        self.ui.label_4.setText(formatted_text)


    def load_table_data(self, file_path):
        try:
            df = pd.read_excel(os.path.join(self.save_dir,file_path))  # Read the Excel file into a DataFrame
            if df.empty:
                QMessageBox.warning(self, "Warning", "The file is empty.")
                return

            # Sort and group by 'Shelter Allocated' and 'Level'
            df = df.sort_values(by=["Shelter Assigned", "Level"])

            self.ui.tableWidget.setRowCount(df.shape[0])  # Set number of rows
            self.ui.tableWidget.setColumnCount(4)  # Ensure 3 columns
            self.ui.tableWidget.setHorizontalHeaderLabels(["Community","Individuals" ,"Shelter Assigned", "Level"])  # Set column headers

            self.ui.tableWidget.setColumnWidth(0, 150)
            self.ui.tableWidget.setColumnWidth(1, 100)
            self.ui.tableWidget.setColumnWidth(2, 170)
            self.ui.tableWidget.setColumnWidth(3, 30)

            # Insert data while tracking merging spans
            shelter_merge_map = {}  # Dict to store start row and span for each shelter
            level_merge_map = {}  # Dict to store start row and span for each level within each shelter

            prev_shelter = None
            prev_level = None
            shelter_start = 0
            level_start = 0

            for row in range(df.shape[0]):
                community, pop, shelter, level = df.iloc[row]

                # Insert community
                self.ui.tableWidget.setItem(row, 0, QTableWidgetItem(str(community)))
                self.ui.tableWidget.setItem(row, 1, QTableWidgetItem(str(pop)))

                # Handle shelter column merging
                if shelter != prev_shelter:
                    if prev_shelter is not None:
                        shelter_merge_map[shelter_start] = row - shelter_start
                    shelter_start = row
                    prev_shelter = shelter

                # Handle level column merging (within the same shelter)
                if level != prev_level or shelter != prev_shelter:
                    if prev_level is not None:
                        level_merge_map[level_start] = row - level_start
                    level_start = row
                    prev_level = level

                self.ui.tableWidget.setItem(row, 2, QTableWidgetItem(str(shelter)))
                self.ui.tableWidget.setItem(row, 3, QTableWidgetItem(str(level)))

            # Final merging for last group
            shelter_merge_map[shelter_start] = df.shape[0] - shelter_start
            level_merge_map[level_start] = df.shape[0] - level_start

            # Merge cells in Shelter Allocated column
            for start, span in shelter_merge_map.items():
                self.ui.tableWidget.setSpan(start, 2, span, 1)

            # Merge cells in Level column
            for start, span in level_merge_map.items():
                self.ui.tableWidget.setSpan(start, 3, span, 1)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {str(e)}")

    
    


if __name__ == "__main__":
    import sys
    from PySide6.QtWidgets import QApplication

    app = QApplication(sys.argv)  # Create application instance
    window = ShelterAllocationReport()  # Create an instance of your dialog
    window.show()  # Show the window
    sys.exit(app.exec())  # Start event loop